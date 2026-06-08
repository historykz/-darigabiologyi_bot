"""Экспорт сдач по разделу в Excel: отдельный лист на каждую неделю."""
import io

from openpyxl import Workbook as XlWorkbook
from openpyxl.styles import Font

from database import crud
from services import roles


async def section_xlsx(section_id: int) -> bytes:
    sec = await crud.get_section(section_id)
    students = await crud.get_students(curator_id=sec.curator_id, group_id=sec.group_id)
    subs = await crud.get_submissions(curator_id=sec.curator_id, section_id=section_id)

    # сгруппируем последние сдачи по (неделя, ученик)
    last: dict[tuple[int, int], object] = {}
    for sub in subs:  # desc по дате
        key = (sub.week, sub.student_id)
        last.setdefault(key, sub)

    wb = XlWorkbook()
    wb.remove(wb.active)
    bold = Font(bold=True)

    for week in range(1, sec.weeks + 1):
        ws = wb.create_sheet(f"Неделя {week}")
        ws.append(["Ученик", "Название файла", "Дата сдачи", "Статус"])
        for c in ws[1]:
            c.font = bold
        for st in students:
            sub = last.get((week, st.id))
            name = f"{st.first_name} {st.last_name}".strip()
            if sub:
                status = f"просрочено +{sub.late_by_minutes} мин" if sub.is_late else "вовремя"
                ws.append([name, sub.submitted_name,
                           roles.fmt_absolute(sub.submitted_at_utc), status])
            else:
                ws.append([name, "—", "—", "не сдал"])
        for col, width in zip("ABCD", (28, 26, 22, 22)):
            ws.column_dimensions[col].width = width

    if not wb.sheetnames:
        wb.create_sheet("Пусто")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
