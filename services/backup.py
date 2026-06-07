"""Создание резервной копии: JSON и Excel."""
import io
import json
from datetime import datetime

from openpyxl import Workbook as XlWorkbook
from sqlalchemy import select

from database.models import (
    Deadline,
    Group,
    Student,
    Submission,
    User,
    Workbook,
)
from database.session import async_session
from services import roles


async def collect() -> dict:
    """Собирает все данные в один словарь."""
    async with async_session() as s:
        users = list((await s.scalars(select(User))).all())
        groups = list((await s.scalars(select(Group))).all())
        students = list((await s.scalars(select(Student))).all())
        workbooks = list((await s.scalars(select(Workbook))).all())
        subs = list((await s.scalars(select(Submission))).all())
        deadlines = list((await s.scalars(select(Deadline))).all())

    return {
        "version": "4.0",
        "created_at": datetime.utcnow().isoformat(),
        "users": [
            {"telegram_id": u.telegram_id, "username": u.username,
             "first_name": u.first_name, "last_name": u.last_name, "role": u.role}
            for u in users
        ],
        "groups": [
            {"id": g.id, "curator_id": g.curator_id, "name": g.name,
             "created_at": g.created_at.isoformat() if g.created_at else None}
            for g in groups
        ],
        "students": [
            {"id": st.id, "user_id": st.user_id, "username": st.username,
             "group_id": st.group_id, "curator_id": st.curator_id,
             "first_name": st.first_name, "last_name": st.last_name,
             "is_active": st.is_active,
             "intro_watched": getattr(st, "intro_watched", False),
             "welcomed": getattr(st, "welcomed", False),
             "added_at": st.added_at.isoformat() if st.added_at else None}
            for st in students
        ],
        "workbooks": [
            {"id": w.id, "serial": w.serial, "topic": w.topic, "file_id": w.file_id,
             "created_at": w.created_at.isoformat() if w.created_at else None}
            for w in workbooks
        ],
        "submissions": [
            {"id": sub.id, "student_id": sub.student_id, "type": sub.type,
             "pdf_file_id": sub.pdf_file_id, "submitted_name": sub.submitted_name,
             "submitted_at_utc": sub.submitted_at_utc.isoformat() if sub.submitted_at_utc else None,
             "curator_id": sub.curator_id, "is_late": sub.is_late,
             "late_by_minutes": sub.late_by_minutes,
             "hidden_for_curator": sub.hidden_for_curator,
             "deleted_at": sub.deleted_at.isoformat() if sub.deleted_at else None}
            for sub in subs
        ],
        "deadlines": [
            {"id": d.id, "curator_id": d.curator_id, "group_id": d.group_id,
             "weekday": d.weekday, "deadline_time_utc": d.deadline_time_utc,
             "reminders_enabled": d.reminders_enabled}
            for d in deadlines
        ],
    }


def make_json(data: dict) -> bytes:
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")


def make_excel(data: dict) -> bytes:
    wb = XlWorkbook()

    # Ученики
    ws = wb.active
    ws.title = "Ученики"
    ws.append(["Имя", "Группа", "Куратор", "Дата добавления"])
    gmap = {g["id"]: g["name"] for g in data["groups"]}
    umap = {u["telegram_id"]: (f"{u['first_name'] or ''} {u['last_name'] or ''}".strip()
                               or u["username"] or str(u["telegram_id"]))
            for u in data["users"]}
    for st in data["students"]:
        if not st["is_active"]:
            continue
        name = f"{st['first_name']} {st['last_name']}".strip()
        added = _fmt(st["added_at"])
        ws.append([name, gmap.get(st["group_id"], "—"),
                   umap.get(st["curator_id"], str(st["curator_id"])), added])

    # Сдачи
    ws2 = wb.create_sheet("Сдачи")
    ws2.append(["Имя", "Группа", "Дата", "Статус"])
    smap = {st["id"]: st for st in data["students"]}
    for sub in data["submissions"]:
        if sub["deleted_at"]:
            continue
        st = smap.get(sub["student_id"], {})
        gname = gmap.get(st.get("group_id"), "—")
        status = f"просрочено +{sub['late_by_minutes']} мин" if sub["is_late"] else "вовремя"
        ws2.append([sub["submitted_name"], gname, _fmt(sub["submitted_at_utc"]), status])

    # Тетради
    ws3 = wb.create_sheet("Тетради")
    ws3.append(["Номер", "Тема", "Дата загрузки"])
    for w in data["workbooks"]:
        ws3.append([f"№{w['serial']:03d}", w["topic"], _fmt(w["created_at"])])

    # Группы
    ws4 = wb.create_sheet("Группы")
    ws4.append(["Название", "Куратор", "Кол-во учеников"])
    counts: dict[int, int] = {}
    for st in data["students"]:
        if st["is_active"]:
            counts[st["group_id"]] = counts.get(st["group_id"], 0) + 1
    for g in data["groups"]:
        ws4.append([g["name"], umap.get(g["curator_id"], str(g["curator_id"])),
                    counts.get(g["id"], 0)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt(iso: str | None) -> str:
    if not iso:
        return "—"
    try:
        return roles.fmt_absolute(datetime.fromisoformat(iso))
    except Exception:
        return iso
